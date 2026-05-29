"""
Excel Engine
============
Exports the complete fitness, nutrition, and tracking plans to a formatted
multi-sheet Excel workbook.

Sheets:
  1. Weekly Workout Schedule  — full 9-column exercise table with clickable demo links
  2. Daily Macro & Meal Plan  — meal-by-meal breakdown
  3. Tracking Strategy        — coach tips, milestones, warnings, metrics
  4. Progress Tracker         — blank template for the client to fill in weekly
"""
import pandas as pd
from typing import List, Optional
from src.state import FitnessPlan, NutritionPlan, TrackingStrategy, ProgressUpdate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill(start_color="1E2A4A", end_color="1E2A4A", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="3B4F81", end_color="3B4F81", fill_type="solid")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL     = PatternFill(start_color="F5F7FF", end_color="F5F7FF", fill_type="solid")
ACCENT_FILL  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
WARN_FILL    = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
THIN_BORDER  = Border(
    left=Side(style="thin", color="D0D7E6"),
    right=Side(style="thin", color="D0D7E6"),
    top=Side(style="thin", color="D0D7E6"),
    bottom=Side(style="thin", color="D0D7E6"),
)

def _style_headers(ws, row=1):
    """Apply dark header style to the first row of a worksheet."""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def _autofit_columns(ws, min_width=10, max_width=50):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

# ---------------------------------------------------------------------------
# Sheet 1: Weekly Workout Schedule
# ---------------------------------------------------------------------------

def _write_workout_sheet(writer, fitness_plan: FitnessPlan):
    """Write the full 9-column exercise table with clickable demo hyperlinks."""
    workout_data = []
    for domain, sessions in [
        ("Gym", fitness_plan.gym_sessions),
        ("Yoga", fitness_plan.yoga_sessions),
        ("Calisthenics", fitness_plan.calisthenics_sessions),
    ]:
        for session in (sessions or []):
            for ex in session.exercises:
                workout_data.append({
                    "Domain": domain,
                    "Day": session.day,
                    "Focus": session.focus,
                    "Duration (min)": session.duration_mins,
                    "Exercise": ex.name,
                    "Sets": ex.sets,
                    "Reps": ex.reps,
                    "Rest (sec)": ex.rest_seconds,
                    "Warm-up Sets": ex.warmup_sets,
                    "Tempo": ex.tempo,
                    "Muscles / Goal": ex.muscles_goal,
                    "Notes": ex.notes or "",
                    "Demo URL": ex.demo_url,  # will be turned into hyperlink below
                })

    columns = [
        "Domain", "Day", "Focus", "Duration (min)", "Exercise",
        "Sets", "Reps", "Rest (sec)", "Warm-up Sets", "Tempo",
        "Muscles / Goal", "Notes", "Demo URL"
    ]
    df = pd.DataFrame(workout_data, columns=columns) if workout_data else pd.DataFrame(columns=columns)
    df.to_excel(writer, sheet_name="Weekly Workout Schedule", index=False)

    ws = writer.sheets["Weekly Workout Schedule"]

    # Convert Demo URL column to actual clickable hyperlinks
    demo_col_idx = columns.index("Demo URL") + 1  # 1-indexed
    demo_col_letter = get_column_letter(demo_col_idx)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        url_cell = ws[f"{demo_col_letter}{row_idx}"]
        if url_cell.value and url_cell.value.startswith("http"):
            url = url_cell.value
            url_cell.value = "▶ Watch Demo"
            url_cell.hyperlink = url
            url_cell.font = Font(color="4A6CF7", underline="single", bold=True)

        # Alternate row shading
        if row_idx % 2 == 0:
            for cell in row:
                if cell.column != demo_col_idx:
                    cell.fill = ALT_FILL

    _style_headers(ws)

# ---------------------------------------------------------------------------
# Sheet 2: Daily Macro & Meal Plan
# ---------------------------------------------------------------------------

def _write_nutrition_sheet(writer, nutrition_plan: Optional[NutritionPlan]):
    """Write the daily meal plan with macro breakdown per meal."""
    meal_data = []
    if nutrition_plan and nutrition_plan.daily_meals:
        for meal in nutrition_plan.daily_meals:
            meal_data.append({
                "Meal": meal.meal_name,
                "Description": meal.description,
                "Calories": meal.calories,
                "Protein (g)": meal.protein_g,
                "Carbs (g)": meal.carbs_g,
                "Fats (g)": meal.fats_g,
            })

    columns = ["Meal", "Description", "Calories", "Protein (g)", "Carbs (g)", "Fats (g)"]
    df = pd.DataFrame(meal_data, columns=columns) if meal_data else pd.DataFrame(columns=columns)

    # Add totals row
    if meal_data:
        totals = {
            "Meal": "DAILY TOTAL",
            "Description": "",
            "Calories": sum(m["Calories"] for m in meal_data),
            "Protein (g)": sum(m["Protein (g)"] for m in meal_data),
            "Carbs (g)": sum(m["Carbs (g)"] for m in meal_data),
            "Fats (g)": sum(m["Fats (g)"] for m in meal_data),
        }
        df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

    df.to_excel(writer, sheet_name="Daily Macro & Meal Plan", index=False)
    ws = writer.sheets["Daily Macro & Meal Plan"]

    # Bold + green fill for totals row
    if meal_data:
        totals_row = ws.max_row
        for cell in ws[totals_row]:
            cell.font = Font(bold=True, color="1B5E20")
            cell.fill = ACCENT_FILL

    # Hydration note below
    if nutrition_plan:
        ws.append([])
        ws.append([f"💧 Daily Hydration Target: {nutrition_plan.hydration_target_L}L"])
        hydration_cell = ws.cell(row=ws.max_row, column=1)
        hydration_cell.font = Font(bold=True, color="1565C0", size=11)

    _style_headers(ws)

# ---------------------------------------------------------------------------
# Sheet 3: Tracking Strategy
# ---------------------------------------------------------------------------

def _write_tracking_sheet(writer, tracking_strategy: Optional[TrackingStrategy]):
    """Write the Tracking Coach's recommendations in a categorized table."""
    rows = []

    if tracking_strategy:
        sections = [
            ("📊 Weekly Check-in Metric", tracking_strategy.weekly_checkin_metrics),
            ("💡 Implementation Tip",     tracking_strategy.implementation_tips),
            ("🎯 Milestone Target",        tracking_strategy.milestone_targets),
            ("⚠️ Red Flag Warning",         tracking_strategy.red_flag_warnings),
        ]
        for category, items in sections:
            for item in (items or []):
                rows.append({"Category": category, "Detail": item})

    columns = ["Category", "Detail"]
    df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
    df.to_excel(writer, sheet_name="Tracking Strategy", index=False)

    ws = writer.sheets["Tracking Strategy"]

    # Color-code rows by category
    category_fills = {
        "📊 Weekly Check-in Metric": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "💡 Implementation Tip":     PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "🎯 Milestone Target":        PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
        "⚠️ Red Flag Warning":         PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    }

    for row_idx in range(2, ws.max_row + 1):
        cat_cell = ws.cell(row=row_idx, column=1)
        detail_cell = ws.cell(row=row_idx, column=2)
        fill = category_fills.get(cat_cell.value)
        if fill:
            cat_cell.fill = fill
            detail_cell.fill = fill
            cat_cell.font = Font(bold=True)
        detail_cell.alignment = Alignment(wrap_text=True)

    # Coach notes section at bottom
    if tracking_strategy and tracking_strategy.coach_notes:
        ws.append([])
        ws.append(["🏅 Coach Notes"])
        notes_header = ws.cell(row=ws.max_row, column=1)
        notes_header.font = Font(bold=True, size=12, color="1E2A4A")
        notes_header.fill = PatternFill(start_color="D1D9FF", end_color="D1D9FF", fill_type="solid")

        ws.append([tracking_strategy.coach_notes])
        notes_cell = ws.cell(row=ws.max_row, column=1)
        notes_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[ws.max_row].height = 100
        ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")

    _style_headers(ws)

# ---------------------------------------------------------------------------
# Sheet 4: Progress Tracker Template
# ---------------------------------------------------------------------------

def _write_tracker_sheet(writer):
    """Blank weekly progress tracker for the client to fill in."""
    df = pd.DataFrame(columns=[
        "Week #", "Date", "Weight (kg)", "Gym Sessions Done",
        "Yoga Sessions Done", "Calisthenics Sessions Done",
        "Nutrition Adherence (1-10)", "Sleep Quality (1-10)",
        "Energy Level (1-10)", "Notes / How I Felt"
    ])
    df.to_excel(writer, sheet_name="Progress Tracker", index=False)
    ws = writer.sheets["Progress Tracker"]
    _style_headers(ws)

    # Add 12 empty rows for 12 weeks
    for i in range(1, 13):
        ws.append([f"Week {i}"] + [""] * 9)

# ---------------------------------------------------------------------------
# Main Export Function
# ---------------------------------------------------------------------------

def export_plans_to_excel(
    fitness_plan: FitnessPlan,
    nutrition_plan: Optional[NutritionPlan],
    tracking_strategy: Optional[TrackingStrategy] = None,
    filepath: str = "Final_Client_Plan.xlsx"
) -> str:
    """
    Exports the verified fitness, nutrition, and tracking plans to a
    beautifully formatted 4-sheet Excel workbook.
    """
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        _write_workout_sheet(writer, fitness_plan)
        _write_nutrition_sheet(writer, nutrition_plan)
        _write_tracking_sheet(writer, tracking_strategy)
        _write_tracker_sheet(writer)

    # Post-process: autofit all columns
    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _autofit_columns(ws)
        ws.freeze_panes = "A2"  # Freeze header row on all sheets

    wb.save(filepath)
    return filepath


def ingest_progress_tracker(filepath: str) -> List[ProgressUpdate]:
    """Parses an uploaded progress tracker Excel file into ProgressUpdate objects."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Progress file not found: {filepath}")

    df = pd.read_excel(filepath, sheet_name="Progress Tracker")
    updates = []
    for _, row in df.iterrows():
        if pd.notna(row.get("Date")) and pd.notna(row.get("Weight (kg)")):
            updates.append(ProgressUpdate(
                date=str(row["Date"]),
                weight_kg=float(row["Weight (kg)"]),
                adherence_score=int(row.get("Nutrition Adherence (1-10)", 5)),
                notes=str(row.get("Notes / How I Felt", ""))
            ))
    return updates
